import glob
from math import ceil
from os import chdir, getcwd

from openpyxl import load_workbook
from openpyxl.styles import Border, PatternFill, Side
from xlwings import Book

from ebay import Parse
from config import ROOT_DIR


class File:
    
    """
    This class helps to find the required file in the directory.
    -----
    * method find_file() helps to find the desired file 
    in a specific directory
    """

    FILENAME: str = None

    def find_file(self, folder: str = None,
                  format: str = ".xlsx") -> str:
        """find_format_file_in_folder."""
        file_list: list = []
        chdir(folder)
        for file in glob.glob(f'*{format}'):
            file_full = folder + '/' + file
            file_list.append((file_full, file))

        if len(file_list) == 1:
            return file_list[0]
        return file_list


class Excel:

    """
    This class allows you to work with excel files, 
    and collect / process the information we need for further work.
    -----
    * method load() read book

    * method check_file() allows you to work and find the desired page in the book, 
      as well as collect information from the desired columns

    * method atribute() forms a dictionary with the necessary columns,
      and removes unnecessary information

    * method find_item() generates a correct dictionary with part numbers

    * method filterkey() removes all extra characters

    * method exceptions() handling exceptions and generating a list of party members
    """

    # collision
    COLLISION: tuple = ('NONE', None, '-', '', ' ', 0)

    # filter_collision
    FILTER_COL: tuple = ('АРТИКУЛ', 'PN', 'МОДЕЛЬ НАЧИНАЕТСЯ С…', 'PART #')

    # sheet_name_collision
    SHEET_COL: tuple = ('Для архива', 'Оценка рыночной стоимости')

    def load(self, dir):
        """Load_book."""
        try:
            book = load_workbook(dir, data_only=False)
            return book
        except:
            return False

    def check_file(self, root_dir, data):
        """Check_sheet_in_input_file."""

        book = self.load(root_dir)
        sheet_list: list = []
        if book:
            sheetname = book.worksheets
            for sheet in sheetname:
                if sheet.title in self.SHEET_COL:
                    continue

                iter_col_dict: dict = {}

                for COL in sheet.iter_cols(1, sheet.max_column):
                    for cell in COL:
                        value = cell.value
                        if type(value) is str:
                            print(value)
                            ind = COL.index(cell) + 1
                            value = value.upper()
                            if value in data:
                                iter_col_dict[value] = COL[ind:]
                                break

                long = len(iter_col_dict)
                if long <= 4:
                    continue
                iter_col_dict['SHEET'] = sheet
                sheet_list.append(iter_col_dict)

        return sheet_list

    def atribute(self, item, ind, data):
        """Create_dict_item."""

        items: dict = {}
        for key in item:
            # if 'МОДУЛИ' in item:
            #     if len(items) == len(item) - 1:
            #         break
            # else
            if len(items) == len(item) - 1:
                break
            if key != 'P/N':
                if key == 'SHEET':
                    continue
                try:
                    value = self.filterkey(item[key][ind].value, key)
                except:
                    continue
                if value != 'None':
                    if 'TEXT' in data[key]:
                        items[key] = str(value)
                    else:
                        items[key] = value
            else:
                items[key] = str(item[key][ind].value)
                    
        return items

    def find_item(self, data, root_dir,):
        """Find_item_in_book."""

        item = self.check_file(root_dir, data)
        if not item:
            return
        final_list: list = []
        for keys in item:
            count: int = 0
            items: dict = {}

            for key in keys['P/N']:
                ind = keys['P/N'].index(key) 
                value = key.value
                if not value:
                    count += 1
                    continue
                filter_value = self.filterkey(value, 'PN')
                if filter_value != 'None':
                    value = str(value).upper()
                    if value not in items:
                        items[value] = self.atribute(keys, ind, data)
                    else:
                        sums = keys['КОЛИЧЕСТВО'][ind].value
                        if sums and type(sums) != str:
                                items[value]['КОЛИЧЕСТВО'] = items[value]['КОЛИЧЕСТВО'] + keys['КОЛИЧЕСТВО'][ind].value
                                if 'COLOR' not in items[value]:
                                    items[value]['COLOR'] = 'YELLOW'

            if count == len(keys['P/N']):
                continue
            items['SHEET'] = str(keys['SHEET'])
            final_list.append(items)
        return final_list


    def filterkey(self, key, col):
        """FilterKeys func."""

        if col == 'КОЛИЧЕСТВО':
            if type(key) is str:
                return 0
            if key in self.COLLISION:
                return 0

        if key in self.COLLISION:
            return 'None'

        if col in self.FILTER_COL:
            key = str(key).replace(' ', '')
            getvals = list([val for val in key if val.isalpha() or val.isnumeric()])
            result = "".join(getvals).upper()
            return result

        return key


    def exceptions(self, key, vendor, exception = {'24': '48',
                                                   '48': '24',
                                                   'K7': ('K8', 'K9'),
                                                   'K8': ('K7', 'K9'),
                                                   'K9': ('K7', 'K8')}):
        """Exceptions."""

        item = [key]
        if vendor:
            if vendor == 'HUAWEI':
                pars = Parse(key).find()
                if pars:
                    item.extend(pars)

            elif vendor == 'CISCO' and 'R-' in key:
                item.append(key.replace('R-', ''))

        else:
            for exc in exception:
                if exc in key:
                    if type(exception[exc]) is tuple:
                        item.append(key.replace(exc, exception[exc][0]))
                        item.append(key.replace(exc, exception[exc][1]))
                        continue
                    item.append(key.replace(exc, exception[exc]))

        return item
    
class Writer(File):

    """
    This class helps to correctly generate the output excel file,
    and record the processed information
    -----
    * method writeinfo() helps to write processed information to excel file

    * method setconst() extract constants from excel
    """

    # set_const
    COUNT: int = 0

    CALCULATION_COLUMN: dict = {'PRICE/USD': '=IF(T{row}="","",T{row}*2+S{row})', 
                                'СТОИМОСТЬ ДОСТАВКИ/USD': '=IF(T{row}="","",T{row}/2)', 
                                'СТ-ТЬ ЗИП С НУЛЯ*1,15': '=IF(T{row}="","",(T{row}*2+S{row})*1.15)',
                                '10% ОТ РЫН.ЦЕНЫ': '=IF(U{row}="","",N{row}*E{row}*0.1)'}

    # Style
    THIN_BORDER = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    THIN_COLOR = PatternFill(start_color='FFFF00', 
                             end_color='FFFF00', 
                             fill_type='solid')

    # root_dir
    OUTPUT_FILE: str =  ROOT_DIR + '\Исходящий\end.xlsx'
    SAMPLE: str = ROOT_DIR + '\Шаблон'

    def __init__(self):
        self.file = self.find_file(self.SAMPLE)

    def writeinfo(self, data):
        
        """Write_info."""

        sample_book = load_workbook(self.file[0], data_only=False)
        ws = sample_book['Расчет']
        ws_archive = sample_book['Для архива']

        point_row = 1
        for key in data:
            self.COUNT += 1
            if key == 'SHEET':
                break
            point_row += 1
            for COL in ws.iter_cols(1, ws.max_column):

                if type(COL[0].value) is not str:
                    continue

                column_val = COL[0].value.upper()
                col_row = COL[0].column
                cell = ws.cell(row=point_row, column=col_row)
                cell.border = self.THIN_BORDER
                if column_val in data[key]:
                    if column_val == 'P/N':
                        if 'COLOR' in data[key]:
                            cell.fill = self.THIN_COLOR
                    if data[key][column_val] == 'None':
                        continue
                    cell.value = data[key][column_val]
                else:
                    if column_val in self.CALCULATION_COLUMN:
                        cell.value = self.CALCULATION_COLUMN[column_val].format(row=point_row)
                        cell.number_format = '0'

        number = 1
        for COL in ws_archive.iter_cols(1, ws_archive.max_column):

            if type(COL[0].value) is not str:
                continue

            value = COL[1].value
            column_val = COL[1].column
            for i in range(2, len(data) + 1):
                number += 1
                cell = ws_archive.cell(row=i, column=column_val)
                cell.border = self.THIN_BORDER
                if value:
                    cell.value = value
                    value = value.replace(str(i), str(i + 1))

        sample_book.save(self.OUTPUT_FILE)

